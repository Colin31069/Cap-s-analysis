from __future__ import annotations

import json
import math
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from skin_analysis.analysis import process_single_file, read_xlsx_single


PARITY_CASES_PATH = Path(__file__).resolve().parents[2] / "shared" / "parity_cases.json"


def load_xlsx_cases() -> list[dict]:
    data = json.loads(PARITY_CASES_PATH.read_text(encoding="utf-8"))
    return data["xlsxCases"]


def expand_segments(segments: list[dict]) -> list[float]:
    values: list[float] = []
    for segment in segments:
        values.extend([float(segment["value"])] * int(segment["repeat"]))
    return values


def write_case_xlsx(path: Path, column_name: str, values: list[float]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.cell(row=1, column=1, value=column_name)
    for row_index, value in enumerate(values, start=2):
        sheet.cell(row=row_index, column=1, value=value)
    workbook.save(path)


class SharedParityCaseTests(unittest.TestCase):
    def test_xlsx_cases_match_shared_expectations(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)

            for case in load_xlsx_cases():
                file_path = tmp_path / f"{case['name']}.xlsx"
                values = expand_segments(case["segments"])
                write_case_xlsx(file_path, case["column"], values)

                expected_error = case.get("expectedError")
                if expected_error == "missing_column":
                    self.assertIsNone(read_xlsx_single(str(file_path)), msg=case["name"])
                    self.assertIsNone(process_single_file(str(file_path)), msg=case["name"])
                    continue

                signal = process_single_file(str(file_path))
                self.assertIsNotNone(signal, msg=case["name"])
                assert signal is not None

                expected = case["expected"]
                self.assertAlmostEqual(signal.initial_avg, expected["initialAvg"], msg=case["name"])
                self.assertAlmostEqual(signal.drop_time, expected["dropTime"], msg=case["name"])

                expected_delta = expected["deltaCapacitance"]
                if expected_delta == "NaN":
                    self.assertTrue(math.isnan(signal.delta_capacitance), msg=case["name"])
                else:
                    self.assertAlmostEqual(signal.delta_capacitance, expected_delta, msg=case["name"])


if __name__ == "__main__":
    unittest.main()
